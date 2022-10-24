#%%
import pandas as pd
import random
import regex as re
from bs4 import BeautifulSoup
from time import sleep
from selenium import webdriver
from openpyxl import load_workbook

def find_elements(html, tag, attr, val):
    soup = BeautifulSoup(html, 'html.parser')
    return soup.find(tag, attrs={attr:val})

def clean_tickers(tick_col):
    return [s.replace('.', ',') if '.' in s else s for s in tick_col]

def month_to_int(date):
    months = {'January':1,
              'February':2,
              'March':3,
              'April':4,
              'May':5,
              'June':6,
              'July':7,
              'August':8,
              'September':9,
              'October': 10,
              'November':11,
              'December':12
    }
    date_split = re.findall(r'[^,\s]+', date)
    return [months[i] if not i.isnumeric() else int(i) for i in date_split]

def largest_str_len(col):
    largest = 0
    for i in col:
        if len(i) > largest:
            largest = len(i)
            val = i
    return len(val.split(','))

def back_month(dt_date, back=1):
    day = dt_date.day
    month = dt_date.month
    if month > back:
        month -= back
        return month, day, dt_date.year
    else:
        return (month - back + 12), day, dt_date.year - 1

def get_sheet_names_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

def make_date_range(date_col, val_indx, frwd_shift=2, back_shift=3):
    if frwd_shift > back_shift:
        raise ValueError('Shift is too large. It must be <= back_shift')

    date = pd.to_datetime(date_col)[val_indx]
    month_min, day, yr = back_month(date, back=back_shift) # goes back
    
    if 10 < month_min < 13:
        month_max = month_min - 12 + frwd_shift # max month is 2 months after min
        yr_min = yr - 1
        return (month_min, day, yr_min), (month_max, day, yr)
    elif month_min == 10:
        month_max = month_min + frwd_shift
        yr_min = yr - 1
        return (month_min, day, yr_min), (month_max, day, yr)
    else:
        month_max = month_min + frwd_shift
        return (month_min, day, yr), (month_max, day, yr)

def make_url(ticker, add_param, m_min, d_min, y_min, m_max, d_max, y_max):
    return f'https://www.google.com/search?q={ticker}{add_param}&rlz=1C1CHBF_enUS1024US1025&biw=1564&bih=932&sxsrf=ALiCzsaGPneyPAo-kyllnxBBtXe-FGWorQ%3A1665448856808&source=lnt&tbs=sbd%3A1%2Ccdr%3A1%2Ccd_min%3A{m_min}%2F{d_min}%2F{y_min}%2Ccd_max%3A{m_max}%2F{d_max}%2F{y_max}&tbm=nws'

#%%
if __name__ == '__main__':
    path = 'C:/Users/agarc/AbbyCode/data/'
    sheet_names = get_sheet_names_xlsx(path+'result_test.xlsx') # exclude jan / feb
    wd = webdriver.Chrome()
    with pd.ExcelWriter(path+"FINAL_result_test.xlsx") as writer:

        for name in sheet_names:
            original = pd.read_excel(path+'result_test.xlsx', sheet_name=name)
            try:
                links = []
                pub_dates = []
                c = 0
                for i, tick in enumerate(original['COMPANY (TICKER)']):
                    date_min, date_max = make_date_range(original['Trade Date'], i)
                    month_min, day_min, yr_min = date_min
                    month_max, day_max, yr_max = date_max

                    url = make_url(tick, '+stock', month_min, day_min, yr_min, month_max, day_max, yr_max)
                    # print(url[181:]) # check dates in url
                    wd.get(url)
                    sleep(random.uniform(1, 4))

                    page = wd.page_source
                    link_elmnt = find_elements(page, tag='a', attr='class', val='WlydOe')
                    try:
                        article_link = link_elmnt.get('href')
                    except (AttributeError) as error:
                        print(error, '\n will append No Results instead of link')
                        links.append('No Results')
                        pub_dates.append(None)
                    else:
                        links.append(article_link) 
                        pub_dates.append(link_elmnt.find('div', attrs={'class':'OSrXXb ZE0LJd YsWzw'}).text)    
                        
            finally:
                wd.quit()

            original['Links'] = links
            original['Link Publish Date'] = pub_dates
            original.to_excel(writer, sheet_name=name, index=False)